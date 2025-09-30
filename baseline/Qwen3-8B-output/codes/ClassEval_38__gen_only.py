import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        data = self.read_excel(save_file_name)
        if not data:
            return (0, save_file_name)
        # Adjust column index to be zero-based
        N -= 1
        processed_data = []
        # Process each row
        for row in data:
            new_row = list(row)
            if N < len(new_row):
                new_row[N] = new_row[N].upper() if isinstance(new_row[N], str) else new_row[N]
            processed_data.append(tuple(new_row))
        # Write processed data to new file
        success = self.write_excel(processed_data, save_file_name)
        return (success, save_file_name)